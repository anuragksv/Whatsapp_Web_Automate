from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
import tkinter as tk
import os
from tkinter import filedialog, Text, Label, messagebox, StringVar, Entry
import openpyxl
import time
from tkinter.scrolledtext import ScrolledText
# Selinium Code


DRIVER_PATH = '.\\tkinter\\chromedriver_86.exe'
excelPath = ''
# webdriver.Chrome(DRIVER_PATH)


def start():
    try:
        global excelPath
        global st
        print(st.get("1.0", 'end-1c'))
        if(excelPath == ''):
            Warn('Please select an excel file')
            return
        if(st.get("1.0", 'end-1c') == ''):
            Warn('Enter the msg')
            return
        driver = webdriver.Chrome(DRIVER_PATH)
        workBook = openpyxl.load_workbook(excelPath.name)
        sheet = workBook.active
        rows = sheet.max_row

        for i in range(1, rows):
            cell = sheet.cell(i, 1)
            if(not str(cell.value).isnumeric()):
                continue
            driver.get('https://api.whatsapp.com/send?phone=+91' +
                       str(cell.value))
            button = driver.find_element_by_xpath('//*[@id="action-button"]')
            driver.implicitly_wait(2)
            ActionChains(driver).move_to_element(
                button).click(button).perform()

            driver.find_element_by_xpath(
                '//*[@id="fallback_block"]/div/div/a').click()
            time.sleep(6)
            if(i == 1):
                while(Submit() != "yes"):
                    pass
            message = st.get('1.0', 'end-1c')
            message += '\n'
            textBox = driver.find_element_by_xpath(
                '//*[@id="main"]/footer/div[1]/div[2]/div/div[2]')
            textBox.click()
            textBox.send_keys(message)
            time.sleep(1)
    except Exception:
        print(Exception.args)
        Warn(Exception.args)

    # TKinter Code


def addExcel():
    global excelPath
    excelPath = filedialog.askopenfile(initialdir='/', title="Select File", filetypes=[
                                       ("Excel files", "*.xlsx *.xls *.xlsm"), ('all', '*')])
    print(excelPath)


def Submit():
    return messagebox.askquestion("Conformation", "Did you Login with WhatsApp Web?")


def Warn(s):
    return messagebox.showwarning("Error", (s))


root = tk.Tk()
root.geometry("1080x720")
root.title("Whatsapp Auto")
root.configure(background="#0866ff")
#

frame = tk.Frame(root, bg="white")
frame.place(relwidth=0.95, relheight=0.95, relx=0.02, rely=0.02)
#
wid = Label(frame, text="Select Your Excel Files")
wid.place(relx=0.47, y=10)
openExcel = tk.Button(frame, text="Open File", bg='white',
                      padx=1, pady=1, command=addExcel)
openExcel.place(relx=0.47, y=45)
#
start = tk.Button(frame, text="Start!", bg='white', padx=1,
                  pady=1, state="active", command=start)
Label(frame, text="Type your message here!").place(relx=0.47, y=82)
st = ScrolledText(frame, width=30, height=10)
st.place(relx=0.47, y=105)
start.place(relx=0.5, y=595)
root.mainloop()


#
